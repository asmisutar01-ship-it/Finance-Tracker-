import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

def format_currency(amount, currency="INR", original_amount=None):
    if currency != "INR" and original_amount is not None:
        return f"{original_amount:,.2f} {currency} (INR {amount:,.2f})"
    return f"INR {amount:,.2f}"

def generate_pdf_report(incomes, expenses, summary, month_str, user_name):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"Monthly Financial Report - {month_str}", styles['Title']))
    elements.append(Paragraph(f"Prepared for: {user_name}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary Section
    elements.append(Paragraph("<b>Summary</b>", styles['Heading2']))
    summary_data = [
        ["Total Income", f"INR {summary['total_income']:,.2f}"],
        ["Total Expenses", f"INR {summary['total_expenses']:,.2f}"],
        ["Net Savings", f"INR {summary['net_savings']:,.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[150, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Transactions Table
    elements.append(Paragraph("<b>Transactions</b>", styles['Heading2']))
    
    table_data = [["Date", "Type", "Category/Source", "Amount"]]
    
    for inc in incomes:
        amount_str = format_currency(inc.get('amount', 0), inc.get('original_currency', 'INR'), inc.get('original_amount'))
        table_data.append([inc.get('date', ''), "Income", inc.get('source', ''), amount_str])
        
    for exp in expenses:
        amount_str = format_currency(exp.get('amount', 0), exp.get('original_currency', 'INR'), exp.get('original_amount'))
        table_data.append([exp.get('date', ''), "Expense", exp.get('category', ''), amount_str])

    # Sort transactions by date
    header = table_data[0]
    rows = table_data[1:]
    rows.sort(key=lambda x: x[0])
    
    final_table_data = [header] + rows

    if len(final_table_data) > 1:
        trans_table = Table(final_table_data, colWidths=[80, 80, 150, 200])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
        ]))
        elements.append(trans_table)
    else:
        elements.append(Paragraph("No transactions found for this month.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(incomes, expenses, summary, month_str):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Report_{month_str}"

    # Header Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    
    # Summary
    ws.append(["Monthly Financial Report", month_str])
    ws.append([])
    
    ws.append(["Summary"])
    ws.append(["Total Income", summary['total_income']])
    ws.append(["Total Expenses", summary['total_expenses']])
    ws.append(["Net Savings", summary['net_savings']])
    ws.append([])

    # Transactions Header
    headers = ["Date", "Type", "Category/Source", "Amount (INR)", "Original Amount", "Original Currency"]
    ws.append(headers)
    
    # Style transaction headers
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Combine and sort transactions
    transactions = []
    for inc in incomes:
        transactions.append({
            'date': inc.get('date', ''),
            'type': 'Income',
            'category': inc.get('source', ''),
            'amount_inr': inc.get('amount', 0),
            'original_amount': inc.get('original_amount', inc.get('amount', 0)),
            'original_currency': inc.get('original_currency', 'INR')
        })
    for exp in expenses:
        transactions.append({
            'date': exp.get('date', ''),
            'type': 'Expense',
            'category': exp.get('category', ''),
            'amount_inr': exp.get('amount', 0),
            'original_amount': exp.get('original_amount', exp.get('amount', 0)),
            'original_currency': exp.get('original_currency', 'INR')
        })
        
    transactions.sort(key=lambda x: x['date'])

    for t in transactions:
        ws.append([
            t['date'],
            t['type'],
            t['category'],
            t['amount_inr'],
            t['original_amount'],
            t['original_currency']
        ])

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
